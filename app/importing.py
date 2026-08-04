import csv
import io
import re

from openpyxl import load_workbook


HEADER_ALIASES = {
    "employee_number": {"payroll number", "payroll numbers", "pay roll number", "pay roll numbers", "payroll no", "payroll id", "employee id", "user id", "staff id"},
    "full_name": {"name", "names", "full name", "staff name", "employee name"},
    "phone": {"phone", "phone number", "phone numbers", "mobile", "mobile number", "telephone"},
    "status": {"status", "duty status", "staff status"},
    "residence": {"residence", "residential area", "estate", "area"},
    "role": {"role", "designation", "job title"},
    "email": {"email", "email address"},
}


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())).strip()


def canonical_header(value: object) -> str | None:
    normalized = normalize_header(value)
    exact = next((field for field, aliases in HEADER_ALIASES.items() if normalized == field.replace("_", " ") or normalized in aliases), None)
    if exact:
        return exact
    words = set(normalized.split())
    if "payroll" in words or {"pay", "roll"}.issubset(words) or ({"employee", "id"}.issubset(words)) or ({"user", "id"}.issubset(words)):
        return "employee_number"
    if "name" in words or "names" in words:
        return "full_name"
    if words.intersection({"phone", "mobile", "telephone", "tel", "contact"}):
        return "phone"
    if "status" in words:
        return "status"
    if words.intersection({"residence", "residential", "estate", "village"}) or normalized in {"home area", "area of residence"}:
        return "residence"
    if words.intersection({"role", "designation"}) or normalized == "job title":
        return "role"
    if "email" in words:
        return "email"
    return None


def canonical_headers(values: list[object]) -> list[str | None]:
    return [canonical_header(value) for value in values]


def locate_header_row(raw_rows: list[list[object] | tuple[object, ...]]) -> tuple[int, list[str | None]]:
    required = {"employee_number", "full_name", "phone", "status", "residence"}
    best_index, best_headers, best_score = 0, [], -1
    for index, row in enumerate(raw_rows[:25]):
        headers = canonical_headers(list(row))
        recognized = {header for header in headers if header}
        score = len(required.intersection(recognized))
        if score > best_score:
            best_index, best_headers, best_score = index, headers, score
        if required.issubset(recognized):
            return index, headers
    recognized = {header for header in best_headers if header}
    missing = ", ".join(sorted(required - recognized))
    found = ", ".join(sorted(recognized)) or "none"
    raise ValueError(f"Could not locate the roster header row. Recognised: {found}. Missing: {missing}")


def text_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_roster(content: bytes, filename: str) -> list[dict[str, str]]:
    if filename.lower().endswith(".xlsx"):
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("Excel file is damaged or is not a valid .xlsx workbook") from exc
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            return []
        header_index, headers = locate_header_row(raw_rows)
        rows = [{header: text_value(value) for header, value in zip(headers, row) if header} for row in raw_rows[header_index + 1:]]
        workbook.close()
    elif filename.lower().endswith(".csv"):
        decoded = content.decode("utf-8-sig")
        raw_rows = list(csv.reader(decoded.splitlines()))
        if not raw_rows:
            return []
        header_index, headers = locate_header_row(raw_rows)
        rows = [{header: text_value(value) for header, value in zip(headers, row) if header} for row in raw_rows[header_index + 1:]]
    else:
        raise ValueError("Roster must be an Excel .xlsx or CSV file")
    return [row for row in rows if any(row.values())]


def normalize_roster_status(value: str) -> str:
    normalized = normalize_header(value)
    if normalized in {"on duty", "duty", "active", "present"}:
        return "on_duty"
    if normalized in {"annual leave", "on leave", "leave"}:
        return "annual_leave"
    raise ValueError(f"Unsupported staff status: {value}")
